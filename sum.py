#!/usr/bin/python3

import openpyxl
import os,glob
from openpyxl import Workbook,load_workbook

cwd = os.getcwd()

wb = Workbook()
ws = wb.active
ws.title = "Summary"

ws['B1'] = 'June'

sheets = glob.glob("*2017.xlsx")
sheet = iter(sheets)
ch_iter = 2
for i in sheet:
    wbi = load_workbook(i)
    wsi = wbi.active
    channel_name = i.split("_201",1)[0]
    current_budget = wsi['B2'].value
    ws.cell(row=ch_iter,column=1,value=channel_name)
    ws.cell(row=ch_iter,column=2,value=current_budget)
    print(channel_name)
    ch_iter+=1

print("Hello World!", sheets)
wb.save("Sum_AAAE.xlsx")
